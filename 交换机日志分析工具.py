import os
import re
import chardet
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from tkinter.font import Font
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pickle


class SwitchLogAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("交换机日志分析工具 v1.0")
        self.root.geometry("1200x800")

        # 设置样式
        self.setup_styles()

        # 创建界面
        self.create_widgets()

        # 存储匹配结果
        self.matching_results = {}
        self.all_results_list = []  # 用于存储所有结果的列表
        self.selected_descriptions = []  # 存储选中的描述，用于Excel列头
        self.current_selected_item = None  # 当前选中的自定义关键词项

    def setup_styles(self):
        """设置界面样式"""
        style = ttk.Style()
        style.theme_use('clam')

        # 配置颜色
        self.colors = {
            'bg': '#f0f0f0',
            'fg': '#333333',
            'select': '#0078d7',
            'success': '#28a745',
            'warning': '#ffc107',
            'error': '#dc3545',
            'info': '#17a2b8'
        }

        # 配置字体
        self.fonts = {
            'title': Font(family="Microsoft YaHei", size=14, weight="bold"),
            'normal': Font(family="Microsoft YaHei", size=10),
            'small': Font(family="Microsoft YaHei", size=9),
            'mono': Font(family="Consolas", size=10)
        }

    def create_widgets(self):
        """创建界面组件"""
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 创建顶部工具栏
        self.create_toolbar(main_frame)

        # 创建主内容区域（左右分栏）
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=10)

        # 左侧控制面板 - 自定义关键词管理
        left_frame = ttk.Frame(paned, width=450)
        paned.add(left_frame, weight=1)
        self.create_keyword_panel(left_frame)

        # 右侧结果显示面板
        right_frame = ttk.Frame(paned, width=750)
        paned.add(right_frame, weight=2)
        self.create_result_panel(right_frame)

        # 创建底部状态栏
        self.create_status_bar(main_frame)

    def create_toolbar(self, parent):
        """创建工具栏"""
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=tk.X, pady=(0, 10))

        # 文件夹选择
        ttk.Label(toolbar, text="日志文件夹:", font=self.fonts['normal']).pack(side=tk.LEFT, padx=(0, 5))
        self.folder_path_var = tk.StringVar(value=r"")
        folder_entry = ttk.Entry(toolbar, textvariable=self.folder_path_var, width=40, font=self.fonts['normal'])
        folder_entry.pack(side=tk.LEFT, padx=(0, 5))

        ttk.Button(toolbar, text="浏览...", command=self.browse_folder, width=8).pack(side=tk.LEFT, padx=(0, 5))

        # 操作按钮
        ttk.Button(toolbar, text="开始分析", command=self.start_analysis, width=8).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar, text="导出Excel", command=self.export_to_excel, width=8).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar, text="清空结果", command=self.clear_results, width=8).pack(side=tk.LEFT)

    def create_keyword_panel(self, parent):
        """创建自定义关键词管理面板"""
        # 标题
        title_label = ttk.Label(parent, text="自定义关键词管理", font=self.fonts['title'])
        title_label.pack(anchor=tk.W, pady=(0, 10))

        # 关键词列表框架
        list_frame = ttk.LabelFrame(parent, text="关键词列表", padding="5")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 创建表格
        columns = ('keyword', 'description', 'before', 'after')
        self.keyword_tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=12)

        self.keyword_tree.heading('keyword', text='关键词')
        self.keyword_tree.heading('description', text='描述')
        self.keyword_tree.heading('before', text='向前提取')
        self.keyword_tree.heading('after', text='向后提取')

        self.keyword_tree.column('keyword', width=120)
        self.keyword_tree.column('description', width=120)
        self.keyword_tree.column('before', width=60)
        self.keyword_tree.column('after', width=60)

        self.keyword_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # 添加滚动条
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.keyword_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.keyword_tree.configure(yscrollcommand=scrollbar.set)

        # 绑定选择事件
        self.keyword_tree.bind('<<TreeviewSelect>>', self.on_keyword_select)

        # 编辑框架
        edit_frame = ttk.LabelFrame(parent, text="编辑关键词", padding="5")
        edit_frame.pack(fill=tk.X, pady=(0, 10))

        # 输入框网格布局
        # 第一行
        ttk.Label(edit_frame, text="关键词:").grid(row=0, column=0, padx=2, pady=2, sticky=tk.W)
        self.keyword_entry = ttk.Entry(edit_frame, width=20)
        self.keyword_entry.grid(row=0, column=1, padx=2, pady=2, sticky=tk.W)

        ttk.Label(edit_frame, text="描述:").grid(row=0, column=2, padx=2, pady=2, sticky=tk.W)
        self.desc_entry = ttk.Entry(edit_frame, width=20)
        self.desc_entry.grid(row=0, column=3, padx=2, pady=2, sticky=tk.W)

        # 第二行
        ttk.Label(edit_frame, text="向前提取(字符):").grid(row=1, column=0, padx=2, pady=2, sticky=tk.W)
        self.before_entry = ttk.Entry(edit_frame, width=10)
        self.before_entry.grid(row=1, column=1, padx=2, pady=2, sticky=tk.W)
        self.before_entry.insert(0, "0")

        ttk.Label(edit_frame, text="向后提取(字符):").grid(row=1, column=2, padx=2, pady=2, sticky=tk.W)
        self.after_entry = ttk.Entry(edit_frame, width=10)
        self.after_entry.grid(row=1, column=3, padx=2, pady=2, sticky=tk.W)
        self.after_entry.insert(0, "50")

        # 按钮框架
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X, pady=5)

        # 第一行按钮
        btn_frame1 = ttk.Frame(btn_frame)
        btn_frame1.pack(fill=tk.X, pady=2)

        ttk.Button(btn_frame1, text="保存", command=self.save_keyword, width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame1, text="删除", command=self.delete_keyword, width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame1, text="清空输入", command=self.clear_input, width=10).pack(side=tk.LEFT, padx=2)

        # 第二行按钮
        btn_frame2 = ttk.Frame(btn_frame)
        btn_frame2.pack(fill=tk.X, pady=2)

        ttk.Button(btn_frame2, text="导出关键词", command=self.export_keywords, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame2, text="导入关键词", command=self.import_keywords, width=12).pack(side=tk.LEFT, padx=2)

        # 全选/取消全选按钮
        select_frame = ttk.Frame(parent)
        select_frame.pack(fill=tk.X, pady=5)

        ttk.Button(select_frame, text="全选所有", command=self.select_all_keywords, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(select_frame, text="取消全选", command=self.deselect_all_keywords, width=12).pack(side=tk.LEFT,
                                                                                                     padx=2)

        # 初始化一些示例关键词
        self.init_sample_keywords()

    def init_sample_keywords(self):
        """初始化示例关键词"""
        sample_keywords = [
            {"keyword": "Software", "description": "软件版本", "before": 0, "after": 50},
            {"keyword": "BER :", "description": "序列号", "before": 0, "after": 30},
            {"keyword": "Uptime is", "description": "运行时间", "before": 0, "after": 40},
            {"keyword": "sysname", "description": "设备名称", "before": 10, "after": 30},
            {"keyword": "Fan", "description": "风扇状态", "before": 0, "after": 60},
            {"keyword": "PowerID", "description": "电源状态", "before": 0, "after": 180},
            {"keyword": "Temperature", "description": "温度信息", "before": 0, "after": 153},
            {"keyword": "-/+ Buffers", "description": "内存使用", "before": 7, "after": 50},
            {"keyword": "CPU utilization", "description": "CPU使用率", "before": 0, "after": 100}
        ]

        for kw in sample_keywords:
            self.keyword_tree.insert('', 'end', values=(
                kw['keyword'], kw['description'], kw['before'], kw['after']
            ))

    def create_result_panel(self, parent):
        """创建结果显示面板"""
        # 创建Notebook
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True)

        # 结果表格页
        table_frame = ttk.Frame(notebook)
        notebook.add(table_frame, text="分析结果")
        self.create_result_table(table_frame)

        # 详细信息页
        detail_frame = ttk.Frame(notebook)
        notebook.add(detail_frame, text="详细信息")
        self.create_detail_view(detail_frame)

        # 统计摘要页
        summary_frame = ttk.Frame(notebook)
        notebook.add(summary_frame, text="统计摘要")
        self.create_summary_view(summary_frame)

    def create_result_table(self, parent):
        """创建结果表格"""
        # 创建Treeview
        columns = ('序号', '文件名', '描述', '提取信息', '状态')
        self.result_tree = ttk.Treeview(parent, columns=columns, show='headings', height=20)

        self.result_tree.heading('序号', text='序号')
        self.result_tree.heading('文件名', text='文件名')
        self.result_tree.heading('描述', text='描述')
        self.result_tree.heading('提取信息', text='提取信息')
        self.result_tree.heading('状态', text='状态')

        self.result_tree.column('序号', width=50)
        self.result_tree.column('文件名', width=200)
        self.result_tree.column('描述', width=150)
        self.result_tree.column('提取信息', width=400)
        self.result_tree.column('状态', width=80)

        # 添加滚动条
        vsb = ttk.Scrollbar(parent, orient="vertical", command=self.result_tree.yview)
        hsb = ttk.Scrollbar(parent, orient="horizontal", command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.result_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # 绑定点击事件
        self.result_tree.bind('<<TreeviewSelect>>', self.on_result_select)

    def create_detail_view(self, parent):
        """创建详细信息视图"""
        # 创建文本框
        self.detail_text = scrolledtext.ScrolledText(
            parent,
            wrap=tk.WORD,
            font=self.fonts['mono'],
            bg='white',
            height=20
        )
        self.detail_text.pack(fill=tk.BOTH, expand=True)

        # 配置标签样式
        self.detail_text.tag_configure('title', font=self.fonts['title'], foreground=self.colors['info'])
        self.detail_text.tag_configure('success', foreground=self.colors['success'])
        self.detail_text.tag_configure('error', foreground=self.colors['error'])
        self.detail_text.tag_configure('warning', foreground=self.colors['warning'])

    def create_summary_view(self, parent):
        """创建统计摘要视图"""
        # 创建统计标签
        self.summary_text = scrolledtext.ScrolledText(
            parent,
            wrap=tk.WORD,
            font=self.fonts['normal'],
            bg='white',
            height=20
        )
        self.summary_text.pack(fill=tk.BOTH, expand=True)

    def create_status_bar(self, parent):
        """创建状态栏"""
        self.status_bar = ttk.Frame(parent, relief=tk.SUNKEN, padding=(2, 2))
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM, pady=(5, 0))

        self.status_label = ttk.Label(self.status_bar, text="就绪", font=self.fonts['small'])
        self.status_label.pack(side=tk.LEFT, padx=5)

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.status_bar, variable=self.progress_var,
                                            length=200, mode='determinate')
        self.progress_bar.pack(side=tk.RIGHT, padx=5)

        self.file_count_label = ttk.Label(self.status_bar, text="文件: 0", font=self.fonts['small'])
        self.file_count_label.pack(side=tk.RIGHT, padx=10)

    def browse_folder(self):
        """浏览文件夹"""
        folder = filedialog.askdirectory()
        if folder:
            self.folder_path_var.set(folder)

    def on_keyword_select(self, event):
        """关键词选择事件"""
        selection = self.keyword_tree.selection()
        if selection:
            self.current_selected_item = selection[0]
            values = self.keyword_tree.item(selection[0])['values']
            if values:
                # 将选中项的值填充到输入框
                self.keyword_entry.delete(0, tk.END)
                self.keyword_entry.insert(0, values[0])

                self.desc_entry.delete(0, tk.END)
                self.desc_entry.insert(0, values[1])

                self.before_entry.delete(0, tk.END)
                self.before_entry.insert(0, str(values[2]))

                self.after_entry.delete(0, tk.END)
                self.after_entry.insert(0, str(values[3]))

    def save_keyword(self):
        """保存关键词（新增或更新）"""
        keyword = self.keyword_entry.get().strip()
        desc = self.desc_entry.get().strip()
        before = self.before_entry.get().strip()
        after = self.after_entry.get().strip()

        if not all([keyword, desc, before, after]):
            messagebox.showwarning("警告", "请填写完整信息")
            return

        try:
            before = int(before)
            after = int(after)

            if self.current_selected_item:
                # 更新现有项
                self.keyword_tree.item(self.current_selected_item, values=(keyword, desc, before, after))
                messagebox.showinfo("成功", "关键词更新成功！")
                self.current_selected_item = None
            else:
                # 新增项
                self.keyword_tree.insert('', 'end', values=(keyword, desc, before, after))
                messagebox.showinfo("成功", "关键词添加成功！")

            self.clear_input()

        except ValueError:
            messagebox.showerror("错误", "向前/向后提取必须为数字")

    def delete_keyword(self):
        """删除关键词"""
        selection = self.keyword_tree.selection()
        if selection:
            if messagebox.askyesno("确认删除", "确定要删除选中的关键词吗？"):
                for item in selection:
                    self.keyword_tree.delete(item)
                self.clear_input()
                self.current_selected_item = None

    def clear_input(self):
        """清空输入框"""
        self.keyword_entry.delete(0, tk.END)
        self.desc_entry.delete(0, tk.END)
        self.before_entry.delete(0, tk.END)
        self.before_entry.insert(0, "0")
        self.after_entry.delete(0, tk.END)
        self.after_entry.insert(0, "50")
        self.current_selected_item = None
        # 取消选择
        self.keyword_tree.selection_remove(self.keyword_tree.selection())

    def export_keywords(self):
        """导出关键词到文件"""
        if not self.keyword_tree.get_children():
            messagebox.showwarning("警告", "没有可导出的关键词！")
            return

        # 选择保存位置
        file_path = filedialog.asksaveasfilename(
            defaultextension=".kw",
            filetypes=[("Keyword files", "*.kw"), ("All files", "*.*")],
            title="导出关键词"
        )

        if file_path:
            try:
                keywords = []
                for item in self.keyword_tree.get_children():
                    values = self.keyword_tree.item(item)['values']
                    keywords.append({
                        'keyword': values[0],
                        'description': values[1],
                        'before': values[2],
                        'after': values[3]
                    })

                with open(file_path, 'wb') as f:
                    pickle.dump(keywords, f)

                messagebox.showinfo("成功", f"关键词已导出到：{file_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{str(e)}")

    def import_keywords(self):
        """从文件导入关键词"""
        file_path = filedialog.askopenfilename(
            filetypes=[("Keyword files", "*.kw"), ("All files", "*.*")],
            title="导入关键词"
        )

        if file_path:
            try:
                with open(file_path, 'rb') as f:
                    keywords = pickle.load(f)

                # 清空现有列表
                for item in self.keyword_tree.get_children():
                    self.keyword_tree.delete(item)

                # 导入新关键词
                for kw in keywords:
                    self.keyword_tree.insert('', 'end', values=(
                        kw['keyword'], kw['description'], kw['before'], kw['after']
                    ))

                messagebox.showinfo("成功", f"成功导入 {len(keywords)} 个关键词！")
            except Exception as e:
                messagebox.showerror("错误", f"导入失败：{str(e)}")

    def select_all_keywords(self):
        """全选所有关键词"""
        for item in self.keyword_tree.get_children():
            self.keyword_tree.selection_add(item)

    def deselect_all_keywords(self):
        """取消全选关键词"""
        self.keyword_tree.selection_remove(self.keyword_tree.selection())

    def get_selected_keywords(self):
        """获取选中的关键词"""
        keywords = []
        self.selected_descriptions = []

        selected_items = self.keyword_tree.selection()
        if not selected_items:
            # 如果没有选中任何项，则使用所有项
            selected_items = self.keyword_tree.get_children()

        for item in selected_items:
            values = self.keyword_tree.item(item)['values']
            if values:
                keywords.append({
                    'keyword': values[0],
                    'before': int(values[2]),
                    'after': int(values[3]),
                    'description': values[1]
                })
                if values[1] not in self.selected_descriptions:
                    self.selected_descriptions.append(values[1])

        return keywords

    def extract_value_without_keyword(self, content, keyword, before, after):
        """提取关键词后的值，不包含关键词本身"""
        start_idx = content.find(keyword)
        if start_idx == -1:
            return None

        # 计算提取范围（从关键词后开始）
        keyword_end = start_idx + len(keyword)
        extract_start = max(0, keyword_end)
        extract_end = min(len(content), keyword_end + after)

        # 提取内容
        extracted = content[extract_start:extract_end].replace('\n', ' ').replace('\r', '')

        # 清理提取的内容
        extracted = extracted.strip()

        # 如果提取内容为空，尝试向后多提取一些
        if not extracted and after < 200:
            extract_end = min(len(content), keyword_end + 200)
            extracted = content[extract_start:extract_end].replace('\n', ' ').replace('\r', '').strip()

        # 限制长度
        extracted = extracted[:200] + ('...' if len(extracted) > 200 else '')

        return extracted

    def start_analysis(self):
        """开始分析"""
        folder_path = self.folder_path_var.get()
        if not os.path.exists(folder_path):
            messagebox.showerror("错误", "文件夹路径不存在！")
            return

        # 清空之前的结果
        self.clear_results()

        # 获取选中的关键词
        keywords = self.get_selected_keywords()
        if not keywords:
            messagebox.showwarning("警告", "请至少选择一个关键词！")
            return

        # 在新线程中执行分析
        thread = threading.Thread(target=self.analyze_files, args=(folder_path, keywords))
        thread.daemon = True
        thread.start()

    def analyze_files(self, folder_path, keywords):
        """分析文件"""
        self.update_status("正在分析文件中...", 0)

        # 获取所有TXT文件
        txt_files = []
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith('.txt'):
                    txt_files.append(os.path.join(root, file))

        total_files = len(txt_files)
        self.file_count_label.config(text=f"文件: {total_files}")

        # 存储结果
        self.all_results_list = []
        result_counter = 1

        for i, file_path in enumerate(txt_files):
            filename = os.path.basename(file_path)
            self.update_status(f"正在分析: {filename}", (i + 1) / total_files * 100)

            try:
                # 检测编码
                with open(file_path, 'rb') as f:
                    raw_data = f.read(10000)
                    result = chardet.detect(raw_data)
                    encoding = result['encoding'] or 'utf-8'

                # 读取文件
                with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
                    content = f.read()

                # 为每个文件创建一个结果字典
                file_result_dict = {'文件名': filename}

                for keyword_info in keywords:
                    keyword = keyword_info['keyword']
                    description = keyword_info['description']
                    before = keyword_info['before']
                    after = keyword_info['after']

                    if keyword in content:
                        # 提取不包含关键词的值
                        extracted = self.extract_value_without_keyword(content, keyword, before, after)
                        if extracted:
                            file_result_dict[description] = extracted

                            # 添加到列表视图
                            self.root.after(0, self.add_to_tree, result_counter, filename,
                                            description, extracted, "成功")
                        else:
                            file_result_dict[description] = "提取失败"
                            self.root.after(0, self.add_to_tree, result_counter, filename,
                                            description, "提取失败", "失败")
                    else:
                        file_result_dict[description] = "未找到"
                        self.root.after(0, self.add_to_tree, result_counter, filename,
                                        description, "未找到", "失败")

                    result_counter += 1

                self.all_results_list.append(file_result_dict)

            except Exception as e:
                error_dict = {'文件名': filename}
                for keyword_info in keywords:
                    error_dict[keyword_info['description']] = f"读取错误: {str(e)}"
                self.all_results_list.append(error_dict)

                self.root.after(0, self.add_to_tree, result_counter, filename,
                                "文件读取错误", str(e), "失败")
                result_counter += 1

        # 更新摘要
        self.root.after(0, self.update_summary)
        self.update_status("分析完成！", 100)

    def add_to_tree(self, seq, filename, description, extracted, status):
        """添加结果到树形视图"""
        self.result_tree.insert('', 'end', values=(seq, filename, description, extracted, status))

    def on_result_select(self, event):
        """结果选择事件"""
        selection = self.result_tree.selection()
        if selection:
            item = self.result_tree.item(selection[0])
            values = item['values']
            if values and len(values) > 1:
                filename = values[1]
                self.show_detail(filename)

    def show_detail(self, filename):
        """显示详细信息"""
        self.detail_text.delete(1.0, tk.END)

        found = False
        for file_result in self.all_results_list:
            if file_result.get('文件名') == filename:
                found = True
                self.detail_text.insert(tk.END, f"文件: {filename}\n", 'title')
                self.detail_text.insert(tk.END, "=" * 50 + "\n\n")

                for key, value in file_result.items():
                    if key != '文件名':
                        self.detail_text.insert(tk.END, f"{key}: ", 'success')
                        self.detail_text.insert(tk.END, f"{value}\n")
                        self.detail_text.insert(tk.END, "-" * 30 + "\n")
                break

        if not found:
            self.detail_text.insert(tk.END, f"未找到文件: {filename}的详细信息")

    def update_summary(self):
        """更新统计摘要"""
        self.summary_text.delete(1.0, tk.END)

        total_files = len(self.all_results_list)
        total_keywords = len(self.selected_descriptions)

        self.summary_text.insert(tk.END, "分析统计摘要\n", 'title')
        self.summary_text.insert(tk.END, "=" * 50 + "\n\n")

        self.summary_text.insert(tk.END, f"分析文件数: {total_files}\n")
        self.summary_text.insert(tk.END, f"关键词数量: {total_keywords}\n\n")

        # 统计每个关键词的匹配情况
        self.summary_text.insert(tk.END, "关键词匹配统计:\n", 'success')
        for desc in self.selected_descriptions:
            match_count = sum(1 for f in self.all_results_list
                              if f.get(desc, '未找到') not in ['未找到', '提取失败'] and '错误' not in f.get(desc, ''))
            self.summary_text.insert(tk.END, f"  {desc}: {match_count}/{total_files} 个文件匹配\n")

    def update_status(self, message, progress):
        """更新状态"""
        self.status_label.config(text=message)
        self.progress_var.set(progress)
        self.root.update_idletasks()

    def export_to_excel(self):
        """导出到Excel文件"""
        if not self.all_results_list:
            messagebox.showwarning("警告", "没有可导出的结果！")
            return

        # 选择保存位置
        output_path = filedialog.askdirectory(title="选择保存位置")
        if not output_path:
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = os.path.join(output_path, f"交换机日志分析结果_{timestamp}.xlsx")

        try:
            # 创建工作簿
            wb = openpyxl.Workbook()

            # 创建主结果表
            ws_main = wb.active
            ws_main.title = "分析结果"

            # 设置表头 - 序号, 文件名, 然后是所有选中的描述
            headers = ['序号', '文件名'] + self.selected_descriptions

            # 设置表头样式
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = ExcelFont(bold=True, color="FFFFFF", size=11)

            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入数据
            for row_idx, file_result in enumerate(self.all_results_list, 2):
                # 序号
                ws_main.cell(row=row_idx, column=1, value=row_idx - 1)
                # 文件名
                ws_main.cell(row=row_idx, column=2, value=file_result.get('文件名', ''))

                # 写入每个描述对应的值
                for col_idx, desc in enumerate(self.selected_descriptions, 3):
                    value = file_result.get(desc, '未找到')
                    cell = ws_main.cell(row=row_idx, column=col_idx, value=value)

                    # 根据值设置背景色
                    if value in ['未找到', '提取失败'] or '错误' in value:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value != '未找到':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            # 调整列宽
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                for row in range(1, len(self.all_results_list) + 2):
                    cell_value = ws_main.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 80)
                ws_main.column_dimensions[column_letter].width = adjusted_width

            # 添加边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in range(1, len(self.all_results_list) + 2):
                for col in range(1, len(headers) + 1):
                    ws_main.cell(row=row, column=col).border = thin_border

            # 创建统计表
            ws_stats = wb.create_sheet(title="统计信息")

            # 统计表头
            stats_headers = ['统计项', '数值']
            for col, header in enumerate(stats_headers, 1):
                cell = ws_stats.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 写入统计数据
            total_files = len(self.all_results_list)
            stats_data = [
                ('分析文件数', total_files),
                ('关键词数量', len(self.selected_descriptions))
            ]

            # 添加每个关键词的统计
            for desc in self.selected_descriptions:
                match_count = sum(1 for f in self.all_results_list
                                  if
                                  f.get(desc, '未找到') not in ['未找到', '提取失败'] and '错误' not in f.get(desc, ''))
                stats_data.append((f'{desc}匹配数', f'{match_count}/{total_files}'))

            for row, (stat, value) in enumerate(stats_data, 2):
                ws_stats.cell(row=row, column=1, value=stat)
                ws_stats.cell(row=row, column=2, value=value)

            # 保存文件
            wb.save(excel_file)

            messagebox.showinfo("成功", f"结果已导出到：\n{excel_file}")

            # 询问是否打开文件
            if messagebox.askyesno("打开文件", "是否打开Excel文件？"):
                os.startfile(excel_file)

        except Exception as e:
            messagebox.showerror("导出失败", f"导出Excel文件时出错：{str(e)}")

    def clear_results(self):
        """清空结果"""
        self.result_tree.delete(*self.result_tree.get_children())
        self.detail_text.delete(1.0, tk.END)
        self.summary_text.delete(1.0, tk.END)
        self.matching_results = {}
        self.all_results_list = []
        self.selected_descriptions = []
        self.progress_var.set(0)
        self.status_label.config(text="就绪")


def main():
    root = tk.Tk()
    app = SwitchLogAnalyzer(root)
    root.mainloop()


if __name__ == "__main__":
    main()